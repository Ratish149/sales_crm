import io
import logging
import os

from django.db.models import Prefetch, Sum
from django.http import FileResponse
from django.utils import timezone
from django_filters import rest_framework as django_filters
from dotenv import load_dotenv
from openpyxl import Workbook
from rest_framework import filters, generics
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from customer.authentication import CustomerJWTAuthentication
from customer.utils import get_customer_from_request
from logistics.models import Logistics
from sales_crm.authentication import TenantJWTAuthentication

from .models import Order, OrderItem
from .serializers import AdminOrderSerializer, OrderListSerializer, OrderSerializer
from .utils import send_order_to_dash

# Optimized queryset for order listings and retrieval
ORDER_OPTIMIZED_QS = (
    Order.objects
    .select_related("customer", "promo_code")
    .prefetch_related(
        Prefetch(
            "items",
            queryset=OrderItem.objects.select_related(
                "product",
                "variant",
                "variant__product",
            ).prefetch_related("variant__option_values__option"),
        )
    )
    .order_by("-created_at")
)

logger = logging.getLogger(__name__)

load_dotenv()
# Reusable function for Dash login
DASH_BASE_URL = os.getenv("DASH_BASE_URL")


class CustomPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class OrderFilter(django_filters.FilterSet):
    status = django_filters.CharFilter(field_name="status", lookup_expr="icontains")
    date_from = django_filters.DateFilter(field_name="created_at", lookup_expr="gte")
    date_to = django_filters.DateFilter(field_name="created_at", lookup_expr="lte")
    is_manual = django_filters.BooleanFilter(field_name="is_manual")
    pos_order = django_filters.BooleanFilter(field_name="pos_order")

    class Meta:
        model = Order
        fields = ["status", "date_from", "date_to", "is_manual", "pos_order"]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        date_from = self.data.get("date_from")
        date_to = self.data.get("date_to")

        # If only date_from is provided, override the filter to get that exact date
        if date_from and not date_to:
            self.filters["date_from"].lookup_expr = "date"  # exact date
            # Optional: remove date_to filter if you want
            if "date_to" in self.filters:
                del self.filters["date_to"]


class OrderListCreateAPIView(generics.ListCreateAPIView):
    queryset = ORDER_OPTIMIZED_QS
    serializer_class = OrderSerializer
    pagination_class = CustomPagination
    authentication_classes = [CustomerJWTAuthentication]
    filter_backends = [
        filters.SearchFilter,
        filters.OrderingFilter,
        django_filters.DjangoFilterBackend,
    ]
    search_fields = ["customer_name", "order_number", "customer_phone"]
    ordering_fields = ["created_at", "total_amount"]
    filterset_class = OrderFilter

    def get_serializer_class(self):
        if self.request.method == "GET":
            return OrderListSerializer
        return OrderSerializer


class AdminOrderListAPIView(generics.ListCreateAPIView):
    queryset = ORDER_OPTIMIZED_QS
    serializer_class = AdminOrderSerializer
    pagination_class = CustomPagination
    filter_backends = [
        filters.SearchFilter,
        filters.OrderingFilter,
        django_filters.DjangoFilterBackend,
    ]
    authentication_classes = [TenantJWTAuthentication]
    permission_classes = [IsAuthenticated]
    search_fields = ["customer_name", "order_number", "customer_phone"]
    ordering_fields = ["created_at", "total_amount"]
    filterset_class = OrderFilter

    def get_serializer_class(self):
        if self.request.method == "GET":
            return OrderListSerializer
        return AdminOrderSerializer


class MyOrderListAPIView(generics.ListAPIView):
    queryset = ORDER_OPTIMIZED_QS
    serializer_class = OrderListSerializer
    pagination_class = CustomPagination
    authentication_classes = [CustomerJWTAuthentication]
    filter_backends = [
        filters.SearchFilter,
        filters.OrderingFilter,
        django_filters.DjangoFilterBackend,
    ]
    search_fields = ["customer_name", "order_number", "customer_phone"]
    ordering_fields = ["created_at", "total_amount"]
    filterset_class = OrderFilter
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        customer = get_customer_from_request(self.request)
        if customer:
            return ORDER_OPTIMIZED_QS.filter(customer=customer)
        return Order.objects.none()


# Retrieve, Update, Delete single Order


class OrderRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ORDER_OPTIMIZED_QS
    serializer_class = OrderSerializer
    authentication_classes = [TenantJWTAuthentication]

    def get_authenticators(self):
        if self.request.method in ["PUT", "PATCH", "DELETE"]:
            return [TenantJWTAuthentication()]
        return []

    def get_permissions(self):
        if self.request.method in ["PUT", "PATCH", "DELETE"]:
            return [IsAuthenticated()]
        return []

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        old_status = instance.status
        response = super().update(request, *args, **kwargs)

        # Refresh instance after update
        instance.refresh_from_db()
        new_status = instance.status

        # Handle transitions to "confirmed" status (Deduct stock and Sync with Logistics)
        if old_status != "confirmed" and new_status == "confirmed":
            instance.deduct_stock()

            # Check if Dash logistics is enabled
            if Logistics.objects.filter(is_enabled=True, logistic="Dash").exists():
                dash_response = send_order_to_dash(instance)

                # If Dash failed, raise error and revert status/stock
                if dash_response.status_code != 200:
                    instance.return_stock()
                    instance.status = old_status
                    instance.save(update_fields=["status"])
                    return dash_response  # Return Dash error directly

        # Handle transitions to "pending" or "cancelled" from a deducted state (Restock)
        DEDUCTED_STATUSES = ["confirmed", "processing", "shipped", "delivered"]
        if old_status in DEDUCTED_STATUSES and new_status in ["pending", "cancelled"]:
            instance.return_stock()

        return response


class OrderGetAPIView(generics.RetrieveAPIView):
    queryset = ORDER_OPTIMIZED_QS
    serializer_class = OrderSerializer
    lookup_field = "order_number"


class OrderExcelExportView(generics.ListAPIView):
    queryset = ORDER_OPTIMIZED_QS
    serializer_class = AdminOrderSerializer
    filter_backends = [
        django_filters.DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = OrderFilter
    authentication_classes = [TenantJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"

        headers = [
            "Order Number",
            "Customer Name",
            "Customer Email",
            "Customer Phone",
            "City",
            "Address",
            "Shipping Address",
            "Payment Type",
            "Status",
            "Total Amount",
            "Delivery Charge",
            "Is Paid",
            "Transaction ID",
            "Note",
            "Created At",
            "Product Name",
            "Variant",
            "Quantity",
            "Unit Price",
            "Item Total",
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        current_row = 2
        for order in queryset:
            # Prefer Customer model details if available
            if order.customer:
                name = f"{order.customer.first_name} {order.customer.last_name}"
                email = order.customer.email
                phone = order.customer.phone
                address = order.customer.address
            else:
                name = order.customer_name
                email = order.customer_email
                phone = order.customer_phone
                address = order.customer_address

            # Base order info
            base_order_info = [
                order.order_number,
                name,
                email,
                phone,
                order.city,
                address,
                order.shipping_address,
                order.get_payment_type_display(),
                order.get_status_display(),
                float(order.total_amount),
                float(order.delivery_charge) if order.delivery_charge else 0.0,
                "Yes" if order.is_paid else "No",
                order.transaction_id,
                order.note,
                (
                    order.created_at.strftime("%Y-%m-%d %H:%M:%S")
                    if order.created_at
                    else ""
                ),
            ]

            items = list(order.items.all())

            for i, item in enumerate(items):
                row_data = [""] * len(headers)

                # Only add order info for the first item row
                if i == 0:
                    row_data[:15] = base_order_info

                # Item info
                product_name = (
                    item.product.name
                    if item.product
                    else (item.variant.product.name if item.variant else "Unknown")
                )
                variant_info = str(item.variant) if item.variant else ""

                row_data[15] = product_name
                row_data[16] = variant_info
                row_data[17] = item.quantity
                row_data[18] = float(item.price)
                row_data[19] = float(item.price * item.quantity)

                for col, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col, value=value)
                current_row += 1

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"orders_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return FileResponse(
            buffer,
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


class DashboardStatsView(APIView):
    authentication_classes = [TenantJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        now = timezone.now()

        total_orders = Order.objects.count()
        total_orders_this_month = Order.objects.filter(
            created_at__year=now.year, created_at__month=now.month
        ).count()

        total_revenue = (
            Order.objects.aggregate(Sum("total_amount"))["total_amount__sum"] or 0
        )

        revenue_this_month = (
            Order.objects.filter(
                created_at__year=now.year, created_at__month=now.month
            ).aggregate(Sum("total_amount"))["total_amount__sum"]
            or 0
        )

        return Response({
            "total_orders": total_orders,
            "total_orders_this_month": total_orders_this_month,
            "total_revenue": total_revenue,
            "revenue_this_month": revenue_this_month,
        })


class MyOrderStatusView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [CustomerJWTAuthentication]

    def get(self, request):
        customer = get_customer_from_request(request)
        if not customer:
            return Response({"error": "Customer not found"}, status=404)

        from django.db.models import Count

        user_orders = Order.objects.filter(customer=customer)
        counts = user_orders.values("status").annotate(count=Count("status"))

        status_counts = {s[0]: 0 for s in Order.ORDER_STATUS}
        for item in counts:
            status_counts[item["status"]] = item["count"]

        return Response(status_counts)
