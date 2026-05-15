import io

from django.http import FileResponse
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated

from sales_crm.authentication import TenantJWTAuthentication
from sales_crm.pagination import CustomPagination
from sales_crm.utils.email_service import get_email_common_context, send_resend_email

from .models import Contact, NewsLetter
from .serializers import (
    ContactListSerializer,
    ContactSerializer,
    NewsLetterListSerializer,
    NewsLetterSerializer,
)

# Optimized querysets for listings
CONTACT_OPTIMIZED_QS = Contact.objects.only(
    "id", "name", "phone_number", "email", "is_read", "created_at"
).order_by("-created_at")

NEWSLETTER_OPTIMIZED_QS = NewsLetter.objects.only(
    "id", "email", "is_subscribed", "is_read", "created_at"
).order_by("-created_at")


class ContactCreateView(generics.ListCreateAPIView):
    queryset = CONTACT_OPTIMIZED_QS
    serializer_class = ContactSerializer
    pagination_class = CustomPagination

    def get_serializer_class(self):
        if self.request.method == "GET":
            return ContactListSerializer
        return ContactSerializer

    def get_authenticators(self):
        # Authenticate only for listing contacts (GET)
        return [TenantJWTAuthentication()] if self.request.method == "GET" else []

    def get_permissions(self):
        # Require authentication only for listing contacts (GET)
        return [IsAuthenticated()] if self.request.method == "GET" else []

    def perform_create(self, serializer):
        contact = serializer.save()

        try:
            context = get_email_common_context()
            context.update({
                "name": contact.name,
                "email": contact.email,
                "phone_number": contact.phone_number,
                "message": contact.message,
            })

            # Admin Notification
            admin_html = render_to_string(
                "contact/email/new_contact_notification.html", context
            )
            send_resend_email(
                context["admin_email"],
                f"New Contact Submission: {contact.name}",
                admin_html,
            )

            # User Acknowledgment
            user_html = render_to_string(
                "contact/email/contact_acknowledgment.html", context
            )
            send_resend_email(
                contact.email,
                f"Thank you for contacting {context['tenant_name']}",
                user_html,
            )

        except Exception as e:
            print(f"Failed to send contact notification emails: {e}")


class ContactRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    authentication_classes = [TenantJWTAuthentication]
    permission_classes = [IsAuthenticated]


class NewsLetterCreateView(generics.ListCreateAPIView):
    queryset = NEWSLETTER_OPTIMIZED_QS
    serializer_class = NewsLetterSerializer
    pagination_class = CustomPagination

    def get_serializer_class(self):
        if self.request.method == "GET":
            return NewsLetterListSerializer
        return NewsLetterSerializer

    def get_authenticators(self):
        # Authenticate only for listing subscribers (GET)
        return [TenantJWTAuthentication()] if self.request.method == "GET" else []

    def get_permissions(self):
        # Require authentication only for listing subscribers (GET)
        return [IsAuthenticated()] if self.request.method == "GET" else []

    def perform_create(self, serializer):
        newsletter = serializer.save()

        try:
            context = get_email_common_context()
            context.update({
                "email": newsletter.email,
            })

            # Admin Notification
            admin_html = render_to_string(
                "contact/email/newsletter_notification.html", context
            )
            send_resend_email(
                context["admin_email"],
                f"New Newsletter Subscription: {newsletter.email}",
                admin_html,
            )

            # User Acknowledgment
            user_html = render_to_string(
                "contact/email/newsletter_acknowledgment.html", context
            )
            send_resend_email(
                newsletter.email,
                f"Thank you for subscribing to {context['tenant_name']}",
                user_html,
            )

        except Exception as e:
            print(f"Failed to send newsletter notification emails: {e}")


class NewsLetterRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = NewsLetter.objects.all()
    serializer_class = NewsLetterSerializer
    authentication_classes = [TenantJWTAuthentication]
    permission_classes = [IsAuthenticated]


class ContactExcelExportView(generics.ListAPIView):
    queryset = CONTACT_OPTIMIZED_QS
    serializer_class = ContactSerializer
    authentication_classes = [TenantJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"

        headers = [
            "Name",
            "Phone Number",
            "Email",
            "Message",
            "Is Read",
            "Created At",
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        current_row = 2
        for contact in queryset:
            row_data = [
                contact.name,
                contact.phone_number,
                contact.email,
                contact.message,
                "Yes" if contact.is_read else "No",
                (
                    contact.created_at.strftime("%Y-%m-%d %H:%M:%S")
                    if contact.created_at
                    else ""
                ),
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"contacts_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return FileResponse(
            buffer,
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
