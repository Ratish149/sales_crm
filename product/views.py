import io
import re
from itertools import chain

import pandas as pd
from django.db.models import (
    Avg,
    Case,
    DecimalField,
    ExpressionWrapper,
    F,
    Max,
    OuterRef,
    Prefetch,
    Subquery,
    Sum,
    When,
)
from django.http import FileResponse
from django_filters import rest_framework as django_filters
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from rest_framework import filters, generics, permissions, serializers, status
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from customer.authentication import CustomerJWTAuthentication
from customer.utils import get_customer_from_request
from sales_crm.authentication import TenantJWTAuthentication
from website.models import SiteConfig

from .models import (
    Category,
    Offer,
    PricingMetric,
    Product,
    ProductComposition,
    ProductImage,
    ProductOption,
    ProductOptionValue,
    ProductReview,
    ProductVariant,
    SubCategory,
    Wishlist,
)
from .serializers import (
    BulkUploadSerializer,
    CategorySerializer,
    OfferSerializer,
    OfferWriteSerializer,
    PricingMetricSerializer,
    ProductCompositionSerializer,
    ProductImageSerializer,
    ProductReviewDetailSerializer,
    ProductReviewSerializer,
    ProductSerializer,
    ProductSmallSerializer,
    ProductVariantAsProductSerializer,
    SubCategoryDetailSerializer,
    SubCategorySerializer,
    UnifiedProductListingSerializer,
    WishlistSerializer,
)
from .utils import (
    download_image_from_url,
    extract_images_from_zip,
    process_image_field,
    safe_value,
)


class CustomPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


class ProductPagination(CustomPagination):
    def get_paginated_response(self, data):
        from collections import OrderedDict

        return Response(
            OrderedDict([
                ("count", self.page.paginator.count),
                ("next", self.get_next_link()),
                ("previous", self.get_previous_link()),
                ("max_price", getattr(self, "max_price", 0.0)),
                ("results", data),
            ])
        )


# ─── final_price annotation helpers ──────────────────────────────────────────
#
# Mirrors the Product.final_price model property at the DB level so we can
# ORDER BY it and compute a correct max_price for the price-range slider.
#
#   use_dynamic_pricing=True  → SUM(metric.price_per_unit × composition.quantity)
#   use_dynamic_pricing=False → product.price

_dynamic_price_sq = (
    ProductComposition.objects
    .filter(product=OuterRef("pk"))
    .values("product")
    .annotate(
        total=Sum(
            ExpressionWrapper(
                F("metric__price_per_unit") * F("quantity"),
                output_field=DecimalField(max_digits=20, decimal_places=2),
            )
        )
    )
    .values("total")[:1]
)

FINAL_PRICE_ANNOTATION = Case(
    When(
        use_dynamic_pricing=True,
        then=Subquery(
            _dynamic_price_sq,
            output_field=DecimalField(max_digits=20, decimal_places=2),
        ),
    ),
    default=F("price"),
    output_field=DecimalField(max_digits=20, decimal_places=2),
)


# ─── Shared optimized queryset ────────────────────────────────────────────────

PRODUCT_LIST_QS = (
    Product.objects
    .select_related("category", "sub_category")
    .prefetch_related(
        Prefetch(
            "images",
            queryset=ProductImage.objects.only("id", "product_id", "image"),
        ),
        Prefetch(
            "variants",
            queryset=ProductVariant.objects.only(
                "id", "product_id", "price", "stock", "image"
            ).prefetch_related(
                Prefetch(
                    "option_values",
                    queryset=ProductOptionValue.objects.only(
                        "id", "value", "option_id"
                    ).select_related("option"),
                )
            ),
        ),
        Prefetch(
            "productoption_set",
            queryset=ProductOption.objects.only(
                "id", "product_id", "name"
            ).prefetch_related(
                Prefetch(
                    "productoptionvalue_set",
                    queryset=ProductOptionValue.objects.only(
                        "id", "option_id", "value"
                    ),
                )
            ),
        ),
        Prefetch(
            "compositions",
            queryset=ProductComposition.objects.only(
                "id", "product_id", "metric_id", "quantity"
            ).select_related("metric"),
        ),
    )
    .annotate(
        average_rating=Avg("productreview__rating"),
        computed_final_price=FINAL_PRICE_ANNOTATION,
    )
    .only(
        "id",
        "name",
        "slug",
        "description",
        "price",
        "market_price",
        "use_dynamic_pricing",
        "base_making_charge",
        "track_stock",
        "stock",
        "weight",
        "thumbnail_image",
        "thumbnail_alt_description",
        "category_id",
        "sub_category_id",
        "is_popular",
        "is_featured",
        "status",
        "fast_shipping",
        "warranty",
        "meta_title",
        "meta_description",
        "created_at",
        "updated_at",
    )
    .order_by("-created_at")
)

CATEGORY_QS = Category.objects.only(
    "id", "name", "slug", "description", "image", "created_at", "updated_at"
)

SUBCATEGORY_QS = SubCategory.objects.select_related("category").only(
    "id",
    "name",
    "slug",
    "description",
    "image",
    "category_id",
    "category__id",
    "category__name",
    "category__slug",
    "created_at",
    "updated_at",
)

PRICING_METRIC_QS = PricingMetric.objects.only(
    "id", "name", "price_per_unit", "unit", "last_updated"
)

PRODUCT_COMPOSITION_QS = ProductComposition.objects.select_related("metric").only(
    "id",
    "product_id",
    "metric_id",
    "quantity",
    "metric__id",
    "metric__name",
    "metric__price_per_unit",
    "metric__unit",
    "metric__last_updated",
)

PRODUCT_IMAGE_QS = ProductImage.objects.only(
    "id", "product_id", "image", "order", "created_at", "updated_at"
)

PRODUCT_REVIEW_QS = (
    ProductReview.objects
    .select_related("product", "user")
    .only("id", "product_id", "user_id", "review", "rating", "created_at", "updated_at")
    .order_by("-created_at")
)

WISHLIST_QS = Wishlist.objects.select_related("user", "product").only(
    "id", "user_id", "product_id", "created_at", "updated_at"
)

PRODUCT_VARIANT_QS = (
    ProductVariant.objects
    .select_related("product", "product__category", "product__sub_category")
    .prefetch_related("option_values__option")
    .only(
        "id",
        "product_id",
        "price",
        "stock",
        "image",
        "created_at",
        "updated_at",
        "product__id",
        "product__name",
        "product__slug",
        "product__price",
        "product__market_price",
        "product__thumbnail_image",
        "product__thumbnail_alt_description",
        "product__is_popular",
        "product__is_featured",
        "product__fast_shipping",
        "product__warranty",
        "product__use_dynamic_pricing",
        "product__base_making_charge",
        "product__created_at",
        "product__updated_at",
        "product__category_id",
        "product__sub_category_id",
        "product__category__id",
        "product__category__name",
        "product__category__slug",
        "product__sub_category__id",
        "product__sub_category__name",
        "product__sub_category__slug",
        "product__sub_category__description",
        "product__sub_category__image",
    )
)


# ─── Category ─────────────────────────────────────────────────────────────────


class CategoryListCreateView(generics.ListCreateAPIView):
    queryset = CATEGORY_QS
    serializer_class = CategorySerializer
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ["name"]


class CategoryRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CATEGORY_QS
    serializer_class = CategorySerializer
    lookup_field = "slug"


# ─── SubCategory ──────────────────────────────────────────────────────────────


class SubCategoryFilterSet(django_filters.FilterSet):
    category = django_filters.CharFilter(
        field_name="category__id", lookup_expr="iexact"
    )

    class Meta:
        model = SubCategory
        fields = ["category"]


class SubCategoryListCreateView(generics.ListCreateAPIView):
    queryset = SUBCATEGORY_QS
    serializer_class = SubCategorySerializer
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter, django_filters.DjangoFilterBackend]
    filterset_class = SubCategoryFilterSet
    search_fields = ["name"]

    def get_serializer_class(self):
        if self.request.method == "POST":
            return SubCategorySerializer
        return SubCategoryDetailSerializer


class SubCategoryRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = SUBCATEGORY_QS
    serializer_class = SubCategorySerializer
    lookup_field = "slug"

    def get_serializer_class(self):
        if self.request.method in ("GET",):
            return SubCategoryDetailSerializer
        return SubCategorySerializer


# ─── Product Image ────────────────────────────────────────────────────────────


class ProductImageListCreateView(generics.ListCreateAPIView):
    queryset = PRODUCT_IMAGE_QS
    serializer_class = ProductImageSerializer

    def get_authenticators(self):
        if self.request.method == "POST":
            return [TenantJWTAuthentication()]
        return []

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated()]
        return super().get_permissions()


class ProductImageRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PRODUCT_IMAGE_QS
    serializer_class = ProductImageSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TenantJWTAuthentication]


# ─── Product filters ──────────────────────────────────────────────────────────


class ProductFilterSet(django_filters.FilterSet):
    category = django_filters.CharFilter(
        field_name="category__slug", lookup_expr="iexact"
    )
    sub_category = django_filters.CharFilter(
        field_name="sub_category__slug", lookup_expr="iexact"
    )
    min_price = django_filters.NumberFilter(field_name="price", lookup_expr="gte")
    max_price = django_filters.NumberFilter(field_name="price", lookup_expr="lte")
    is_popular = django_filters.BooleanFilter(field_name="is_popular")
    is_featured = django_filters.BooleanFilter(field_name="is_featured")
    offer = django_filters.CharFilter(method="filter_by_offer")
    flash_sales = django_filters.BooleanFilter(method="filter_flash_sales")  # ← add

    class Meta:
        model = Product
        fields = [
            "category",
            "sub_category",
            "min_price",
            "max_price",
            "is_popular",
            "is_featured",
            "offer",
            "flash_sales",  # ← add
        ]

    def filter_flash_sales(self, queryset, name, value):
        if not value:
            return queryset

        from django.db.models import Q
        from django.utils import timezone

        now = timezone.now()
        active_offers = Offer.objects.filter(
            is_active=True, start_date__lte=now, end_date__gte=now
        )

        if not active_offers.exists():
            return queryset  # no active offers → return all products

        category_ids = list(
            filter(None, active_offers.values_list("categories", flat=True))
        )
        sub_category_ids = list(
            filter(None, active_offers.values_list("sub_categories", flat=True))
        )
        product_ids = list(
            filter(None, active_offers.values_list("products", flat=True))
        )

        q = Q()
        if category_ids:
            q |= Q(category_id__in=category_ids)
        if sub_category_ids:
            q |= Q(sub_category_id__in=sub_category_ids)
        if product_ids:
            q |= Q(id__in=product_ids)

        if not q:
            return queryset  # offers exist but have no rules → return all

        return queryset.filter(q).distinct()

    def filter_by_offer(self, queryset, name, value):
        if not value:
            return queryset
        try:
            offer = Offer.objects.get(slug=value)
        except Offer.DoesNotExist:
            return queryset.none()

        from django.db.models import Q

        q_objects = Q()
        has_conditions = False

        if offer.products.exists():
            q_objects |= Q(id__in=offer.products.values_list("id", flat=True))
            has_conditions = True

        if offer.categories.exists():
            q_objects |= Q(category__in=offer.categories.all())
            has_conditions = True

        if offer.sub_categories.exists():
            q_objects |= Q(sub_category__in=offer.sub_categories.all())
            has_conditions = True

        if not has_conditions:
            return queryset.none()

        return queryset.filter(q_objects).distinct()


class ProductVariantFilterSet(django_filters.FilterSet):
    min_price = django_filters.NumberFilter(field_name="price", lookup_expr="gte")
    max_price = django_filters.NumberFilter(field_name="price", lookup_expr="lte")
    category = django_filters.CharFilter(field_name="product__category__slug")
    sub_category = django_filters.CharFilter(field_name="product__sub_category__slug")
    status = django_filters.ChoiceFilter(
        field_name="product__status", choices=Product.STATUS_CHOICES
    )
    is_popular = django_filters.BooleanFilter(field_name="product__is_popular")
    is_featured = django_filters.BooleanFilter(field_name="product__is_featured")
    offer = django_filters.CharFilter(method="filter_by_offer")
    flash_sales = django_filters.BooleanFilter(method="filter_flash_sales")  # ← add

    class Meta:
        model = ProductVariant
        fields = [
            "category",
            "sub_category",
            "min_price",
            "max_price",
            "status",
            "is_popular",
            "is_featured",
            "offer",
            "flash_sales",  # ← add
        ]

    def filter_flash_sales(self, queryset, name, value):
        if not value:
            return queryset

        from django.db.models import Q
        from django.utils import timezone

        now = timezone.now()
        active_offers = Offer.objects.filter(
            is_active=True, start_date__lte=now, end_date__gte=now
        )

        if not active_offers.exists():
            return queryset  # no active offers → return all variants

        category_ids = list(
            filter(None, active_offers.values_list("categories", flat=True))
        )
        sub_category_ids = list(
            filter(None, active_offers.values_list("sub_categories", flat=True))
        )
        product_ids = list(
            filter(None, active_offers.values_list("products", flat=True))
        )

        q = Q()
        if category_ids:
            q |= Q(product__category_id__in=category_ids)
        if sub_category_ids:
            q |= Q(product__sub_category_id__in=sub_category_ids)
        if product_ids:
            q |= Q(product_id__in=product_ids)

        if not q:
            return queryset

        return queryset.filter(q).distinct()

    def filter_by_offer(self, queryset, name, value):
        if not value:
            return queryset
        try:
            offer = Offer.objects.get(slug=value)
        except Offer.DoesNotExist:
            return queryset.none()

        from django.db.models import Q

        q_objects = Q()
        has_conditions = False

        if offer.products.exists():
            q_objects |= Q(product_id__in=offer.products.values_list("id", flat=True))
            has_conditions = True

        if offer.categories.exists():
            q_objects |= Q(product__category__in=offer.categories.all())
            has_conditions = True

        if offer.sub_categories.exists():
            q_objects |= Q(product__sub_category__in=offer.sub_categories.all())
            has_conditions = True

        if not has_conditions:
            return queryset.none()

        return queryset.filter(q_objects).distinct()


# ─── Product ──────────────────────────────────────────────────────────────────


class ProductListCreateView(generics.ListCreateAPIView):
    queryset = PRODUCT_LIST_QS
    serializer_class = ProductSerializer
    pagination_class = ProductPagination
    filter_backends = [
        django_filters.DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = ProductFilterSet
    ordering_fields = ["created_at", "price", "is_popular", "average_rating"]
    search_fields = ["name"]

    def get_authenticators(self):
        if self.request.method == "POST":
            return [TenantJWTAuthentication()]
        return []

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_serializer_class(self):
        site_config = SiteConfig.get_solo()
        if self.request.method == "POST":
            return ProductSerializer
        if site_config.use_product_variant:
            return UnifiedProductListingSerializer
        return ProductSmallSerializer

    def get_queryset(self):
        if self.request.method == "GET":
            return PRODUCT_LIST_QS
        return Product.objects.all()

    def filter_queryset(self, queryset):
        """
        Intercept ordering=price / ordering=-price and redirect to the
        computed_final_price annotation so dynamic-priced products sort
        correctly.  All other filters/ordering pass through unchanged.
        """
        queryset = super().filter_queryset(queryset)
        ordering = self.request.query_params.get("ordering", "")
        if ordering.lstrip("-") == "price":
            direction = "-" if ordering.startswith("-") else ""
            queryset = queryset.order_by(f"{direction}computed_final_price")
        return queryset

    def list(self, request, *args, **kwargs):
        site_config = SiteConfig.get_solo()

        if not site_config.use_product_variant:
            queryset = self.filter_queryset(self.get_queryset())

            # Global max_price across ALL products using real final prices,
            # unaffected by filters or pagination
            max_price = float(
                Product.objects.annotate(
                    computed_final_price=FINAL_PRICE_ANNOTATION
                ).aggregate(m=Max("computed_final_price"))["m"]
                or 0.0
            )

            page = self.paginate_queryset(queryset)
            if page is not None:
                if self.paginator is not None:
                    self.paginator.max_price = max_price
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)

        # ── variant mode ────────────────────────────────────────────────────────
        self.filter_backends = []

        params = request.GET.copy()

        from distutils.util import strtobool

        if "is_popular" in params:
            try:
                params["is_popular"] = strtobool(str(params["is_popular"]))
            except ValueError:
                params.pop("is_popular")

        variant_qs = PRODUCT_VARIANT_QS
        variant_filter = ProductVariantFilterSet(params, queryset=variant_qs)
        variant_qs = variant_filter.qs

        search_filter = filters.SearchFilter()
        ordering_filter = filters.OrderingFilter()

        orig_search_fields = self.search_fields
        self.search_fields = ["product__name", "option_values__value"]

        variant_qs = search_filter.filter_queryset(request, variant_qs, self)
        variant_qs = ordering_filter.filter_queryset(request, variant_qs, self)

        self.search_fields = orig_search_fields

        product_qs = (
            Product.objects
            .filter(variants__isnull=True)
            .select_related("category", "sub_category")
            .prefetch_related(
                Prefetch(
                    "images",
                    queryset=ProductImage.objects.only("id", "product_id", "image"),
                ),
            )
            .annotate(
                average_rating=Avg("productreview__rating"),
                computed_final_price=FINAL_PRICE_ANNOTATION,
            )
            .only(
                "id",
                "name",
                "slug",
                "price",
                "market_price",
                "stock",
                "thumbnail_image",
                "thumbnail_alt_description",
                "category_id",
                "sub_category_id",
                "is_popular",
                "is_featured",
                "fast_shipping",
                "warranty",
                "use_dynamic_pricing",
                "base_making_charge",
                "created_at",
                "updated_at",
            )
        )

        product_filter = ProductFilterSet(params, queryset=product_qs)
        product_qs = product_filter.qs
        product_qs = search_filter.filter_queryset(request, product_qs, self)
        product_qs = ordering_filter.filter_queryset(request, product_qs, self)

        # ── ordering-aware Python sort for the combined list ──────────────────
        ordering = request.query_params.get("ordering", "-created_at")
        sort_field = ordering.lstrip("-")
        sort_reverse = ordering.startswith("-")

        def _sort_key(obj):
            if sort_field == "price":
                if isinstance(obj, ProductVariant):
                    # Variants use their own price (no dynamic pricing on variants)
                    return float(obj.price or 0)
                # Products: use the annotated computed_final_price if available,
                # otherwise fall back to the model property
                if (
                    hasattr(obj, "computed_final_price")
                    and obj.computed_final_price is not None
                ):
                    return float(obj.computed_final_price)
                return float(obj.final_price or 0)
            if sort_field == "average_rating":
                return float(getattr(obj, "average_rating", None) or 0)
            if sort_field == "is_popular":
                val = getattr(obj, "is_popular", False)
                if isinstance(obj, ProductVariant):
                    val = getattr(obj.product, "is_popular", False)
                return 0 if val else 1  # True sorts first when reverse=True
            # Default / created_at
            return getattr(obj, "created_at", None) or 0

        combined = sorted(
            chain(variant_qs, product_qs),
            key=_sort_key,
            reverse=sort_reverse,
        )

        # Global max_price across ALL variants and ALL standalone products,
        # unaffected by filters or pagination
        variant_max = float(
            ProductVariant.objects.aggregate(m=Max("price"))["m"] or 0.0
        )
        standalone_product_max = float(
            Product.objects
            .filter(variants__isnull=True)
            .annotate(computed_final_price=FINAL_PRICE_ANNOTATION)
            .aggregate(m=Max("computed_final_price"))["m"]
            or 0.0
        )
        # Also account for variants with no price, which fall back to product.price
        fallback_max = float(
            Product.objects.filter(
                variants__isnull=False,
                variants__price__isnull=True,
            ).aggregate(m=Max("price"))["m"]
            or 0.0
        )
        max_price = max(variant_max, standalone_product_max, fallback_max)

        page = self.paginate_queryset(combined)
        if page is not None:
            if self.paginator is not None:
                self.paginator.max_price = max_price
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(combined, many=True)
        return Response(serializer.data)


class AdminProductListCreateView(generics.ListCreateAPIView):
    queryset = PRODUCT_LIST_QS
    serializer_class = ProductSerializer
    pagination_class = CustomPagination
    filter_backends = [
        django_filters.DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = ProductFilterSet
    ordering_fields = ["created_at", "price", "is_popular", "average_rating"]
    search_fields = ["name"]

    def get_authenticators(self):
        if self.request.method == "POST":
            return [TenantJWTAuthentication()]
        return []

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_serializer_class(self):
        if self.request.method == "POST":
            return ProductSerializer
        return ProductSmallSerializer

    def get_queryset(self):
        if self.request.method == "GET":
            return PRODUCT_LIST_QS
        return Product.objects.all()

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)
        ordering = self.request.query_params.get("ordering", "")
        if ordering.lstrip("-") == "price":
            direction = "-" if ordering.startswith("-") else ""
            queryset = queryset.order_by(f"{direction}computed_final_price")
        return queryset


class ProductRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PRODUCT_LIST_QS
    serializer_class = ProductSerializer
    lookup_field = "slug"

    def get_authenticators(self):
        if self.request.method in ["PUT", "PATCH", "DELETE"]:
            return [TenantJWTAuthentication()]
        return []

    def get_permissions(self):
        if self.request.method in ["PUT", "PATCH", "DELETE"]:
            return [IsAuthenticated()]
        return super().get_permissions()

    def get_queryset(self):
        if self.request.method == "GET":
            return PRODUCT_LIST_QS
        return Product.objects.all()


class RelatedProductList(generics.ListAPIView):
    queryset = PRODUCT_LIST_QS
    serializer_class = ProductSmallSerializer
    filter_backends = [
        django_filters.DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = ProductFilterSet
    ordering_fields = ["created_at", "price", "is_popular", "average_rating"]
    search_fields = ["name"]

    def get_queryset(self):
        try:
            product_slug = self.kwargs["slug"]
            product = Product.objects.get(slug=product_slug)
        except Product.DoesNotExist:
            return Product.objects.none()

        return Product.objects.filter(
            category=product.category,
        ).exclude(id=product.id)[:4]


class ProductExcelExportView(generics.ListAPIView):
    queryset = PRODUCT_LIST_QS
    serializer_class = ProductSerializer
    filter_backends = [
        django_filters.DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = ProductFilterSet
    authentication_classes = [TenantJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        headers = [
            "name",
            "description",
            "price",
            "market_price",
            "track_stock",
            "stock",
            "weight",
            "thumbnail_image",
            "thumbnail_image_aly_description",
            "category",
            "subcategory",
            "is_popular",
            "is_featured",
            "status",
            "use_dynamic_pricing",
            "base_making_charge",
            "require_custom_image",
            "composition",
            "quantity",
            "option1 name",
            "option1 values",
            "option2 name",
            "option2 values",
            "option3 name",
            "option3 values",
            "variant price",
            "variant stock",
            "variant image",
            "meta title",
            "meta description",
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        current_row = 2
        for product in queryset:
            base_data = [
                product.name,
                product.description,
                float(product.price) if product.price else 0.0,
                float(product.market_price) if product.market_price else 0.0,
                "TRUE" if product.track_stock else "FALSE",
                product.stock,
                product.weight,
                product.thumbnail_image.url if product.thumbnail_image else "",
                product.thumbnail_alt_description,
                product.category.name if product.category else "",
                product.sub_category.name if product.sub_category else "",
                "TRUE" if product.is_popular else "FALSE",
                "TRUE" if product.is_featured else "FALSE",
                product.status,
                "TRUE" if product.use_dynamic_pricing else "FALSE",
                float(product.base_making_charge)
                if product.base_making_charge
                else 0.0,
                "TRUE" if product.require_custom_image else "FALSE",
            ]

            compositions = list(product.compositions.all())
            variants = list(product.variants.all())
            options = list(product.productoption_set.all())

            option_headers = ["", "", "", "", "", ""]
            for i, opt in enumerate(options[:3]):
                option_headers[i * 2] = opt.name

            row_data = (
                base_data
                + ["", ""]
                + option_headers
                + ["", "", "", product.meta_title, product.meta_description]
            )

            if compositions:
                comp = compositions[0]
                row_data[17] = (
                    f"{comp.metric.name} ({comp.metric.price_per_unit}/{comp.metric.unit})"
                )
                row_data[18] = float(comp.quantity)

            if variants:
                var = variants[0]
                var_values = list(var.option_values.all())
                for val in var_values:
                    for i, opt in enumerate(options[:3]):
                        if val.option_id == opt.id:
                            row_data[20 + i * 2] = val.value
                row_data[25] = float(var.price) if var.price else 0.0
                row_data[26] = var.stock
                row_data[27] = var.image.url if var.image else ""

            for col, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col, value=value)
            current_row += 1

            max_remaining = max(len(compositions), len(variants))
            for i in range(1, max_remaining):
                extra_row = [""] * len(headers)
                if i < len(compositions):
                    comp = compositions[i]
                    extra_row[17] = (
                        f"{comp.metric.name} ({comp.metric.price_per_unit}/{comp.metric.unit})"
                    )
                    extra_row[18] = float(comp.quantity)

                if i < len(variants):
                    var = variants[i]
                    var_values = list(var.option_values.all())
                    for val in var_values:
                        for j, opt in enumerate(options[:3]):
                            if val.option_id == opt.id:
                                extra_row[20 + j * 2] = val.value
                    extra_row[25] = float(var.price) if var.price else 0.0
                    extra_row[26] = var.stock
                    extra_row[27] = var.image.url if var.image else ""

                for col, value in enumerate(extra_row, 1):
                    ws.cell(row=current_row, column=col, value=value)
                current_row += 1

        column_widths = {
            "A": 15,
            "B": 20,
            "C": 10,
            "D": 15,
            "E": 12,
            "F": 10,
            "G": 10,
            "H": 20,
            "I": 20,
            "J": 15,
            "K": 25,
            "L": 12,
            "M": 12,
            "N": 10,
            "O": 20,
            "P": 20,
            "Q": 20,
            "R": 30,
            "S": 12,
            "T": 15,
            "U": 15,
            "V": 15,
            "W": 15,
            "X": 15,
            "Y": 15,
            "Z": 15,
            "AA": 15,
            "AB": 20,
            "AC": 20,
            "AD": 25,
        }
        for col_let, width in column_widths.items():
            ws.column_dimensions[col_let].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = (
            f"products_export_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return FileResponse(
            buffer,
            as_attachment=True,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ─── Pricing Metric ───────────────────────────────────────────────────────────


class PricingMetricListCreateView(generics.ListCreateAPIView):
    queryset = PRICING_METRIC_QS
    serializer_class = PricingMetricSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TenantJWTAuthentication]


class PricingMetricRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PRICING_METRIC_QS
    serializer_class = PricingMetricSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TenantJWTAuthentication]


# ─── Product Composition ──────────────────────────────────────────────────────


class ProductCompositionListCreateView(generics.ListCreateAPIView):
    queryset = PRODUCT_COMPOSITION_QS
    serializer_class = ProductCompositionSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TenantJWTAuthentication]


class ProductCompositionRetrieveUpdateDestroyView(
    generics.RetrieveUpdateDestroyAPIView
):
    queryset = PRODUCT_COMPOSITION_QS
    serializer_class = ProductCompositionSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TenantJWTAuthentication]


# ─── Product Review ───────────────────────────────────────────────────────────


class ProductReviewFilter(django_filters.FilterSet):
    rating = django_filters.NumberFilter(field_name="rating", lookup_expr="exact")

    class Meta:
        model = ProductReview
        fields = ["rating"]


class ProductReviewView(generics.ListCreateAPIView):
    queryset = PRODUCT_REVIEW_QS
    serializer_class = ProductReviewSerializer
    pagination_class = CustomPagination
    filter_backends = [django_filters.DjangoFilterBackend]
    filterset_class = ProductReviewFilter
    authentication_classes = [CustomerJWTAuthentication]

    def get_permissions(self):
        if self.request.method == "POST":
            return [permissions.IsAuthenticated()]
        return [permissions.AllowAny()]

    def get_serializer_class(self):
        if self.request.method == "GET":
            return ProductReviewDetailSerializer
        return ProductReviewSerializer

    def get_queryset(self):
        slug = self.request.query_params.get("slug")
        try:
            product = Product.objects.only("id").get(slug=slug)
            return PRODUCT_REVIEW_QS.filter(product=product)
        except Product.DoesNotExist:
            return PRODUCT_REVIEW_QS

    def perform_create(self, serializer):
        user = get_customer_from_request(self.request)
        if not user:
            raise serializers.ValidationError(
                "User must be authenticated to create a review."
            )
        serializer.save(user=user)


class ProductReviewRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PRODUCT_REVIEW_QS
    serializer_class = ProductReviewSerializer
    authentication_classes = [CustomerJWTAuthentication]
    lookup_field = "id"

    def get_permissions(self):
        if self.request.method == "GET":
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]


# ─── Wishlist ─────────────────────────────────────────────────────────────────


class WishlistListCreateView(generics.ListCreateAPIView):
    queryset = WISHLIST_QS
    serializer_class = WishlistSerializer
    authentication_classes = [CustomerJWTAuthentication]

    def get_queryset(self):
        user = get_customer_from_request(self.request)
        if not user:
            return Wishlist.objects.none()
        return WISHLIST_QS.filter(user=user)

    def perform_create(self, serializer):
        user = get_customer_from_request(self.request)
        if not user:
            raise serializers.ValidationError(
                "User must be authenticated to add to wishlist."
            )
        serializer.save(user=user)


class WishlistRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = WISHLIST_QS
    serializer_class = WishlistSerializer
    authentication_classes = [CustomerJWTAuthentication]
    lookup_field = "id"

    def get_object(self):
        try:
            user = get_customer_from_request(self.request)
            return Wishlist.objects.get(user=user)
        except Wishlist.DoesNotExist:
            return Response(
                {"message": "Wishlist not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

    def delete(self, request, *args, **kwargs):
        try:
            id = self.kwargs.get("id")
            user = get_customer_from_request(self.request)
            wishlist = Wishlist.objects.get(user=user, id=id)
            wishlist.delete()
            return Response(
                {"message": "Product removed from wishlist"},
                status=status.HTTP_204_NO_CONTENT,
            )
        except Wishlist.DoesNotExist:
            return Response(
                {"message": "Wishlist not found"},
                status=status.HTTP_404_NOT_FOUND,
            )


# ─── Product Variant ──────────────────────────────────────────────────────────


class ProductVariantListCreateView(generics.ListCreateAPIView):
    queryset = PRODUCT_VARIANT_QS
    serializer_class = ProductVariantAsProductSerializer
    pagination_class = CustomPagination
    filter_backends = [
        django_filters.DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = ProductVariantFilterSet
    search_fields = ["product__name", "option_values__value"]
    ordering_fields = ["created_at", "price"]

    def get_authenticators(self):
        if self.request.method == "POST":
            return [TenantJWTAuthentication()]
        return []

    def get_permissions(self):
        if self.request.method == "POST":
            return [IsAuthenticated()]
        return super().get_permissions()


class ProductVariantRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PRODUCT_VARIANT_QS
    serializer_class = ProductVariantAsProductSerializer

    def get_authenticators(self):
        if self.request.method in ["PUT", "PATCH", "DELETE"]:
            return [TenantJWTAuthentication()]
        return []

    def get_permissions(self):
        if self.request.method in ["PUT", "PATCH", "DELETE"]:
            return [IsAuthenticated()]
        return super().get_permissions()


# ─── Download Template ────────────────────────────────────────────────────────


class DownloadProductSampleTemplateView(APIView):
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Template"

        metrics = list(
            PricingMetric.objects.only("id", "name", "price_per_unit", "unit")
        )

        headers = [
            "name",
            "description",
            "price",
            "market_price",
            "track_stock",
            "stock",
            "weight",
            "thumbnail_image",
            "thumbnail_image_aly_description",
            "images",
            "category",
            "subcategory",
            "is_popular",
            "is_featured",
            "status",
            "use_dynamic_pricing",
            "base_making_charge",
            "require_custom_image",
            "composition",
            "quantity",
            "option1 name",
            "option1 values",
            "option2 name",
            "option2 values",
            "option3 name",
            "option3 values",
            "variant price",
            "variant stock",
            "variant image",
            "meta title",
            "meta description",
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        categories = Category.objects.only("id", "name")
        subcategories = SubCategory.objects.select_related("category").only(
            "id", "name", "category__name"
        )

        sample_category = (
            categories.first().name if categories.exists() else "Electronics"
        )
        sample_subcategory = (
            subcategories.first().name if subcategories.exists() else "Mobile"
        )
        sample_metric = (
            f"{metrics[0].name} ({metrics[0].price_per_unit}/{metrics[0].unit})"
            if metrics
            else "Gold (13000.00/gram)"
        )

        sample_data = [
            "productname",
            "product description",
            100,
            100,
            "TRUE",
            15,
            "45g",
            "image.jpg",
            "alt description",
            "image1.jpg, image2.jpg",
            sample_category,
            sample_subcategory,
            "TRUE",
            "FALSE",
            "active",
            "TRUE",
            0.00,
            "FALSE",
            sample_metric,
            5,
            "Color",
            "Green",
            "Size",
            "S",
            "",
            "",
            80,
            10,
            "image.jpg",
            "Meta Title",
            "Meta Description",
        ]

        for col, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col, value=value)

        second_metric = (
            f"{metrics[1].name} ({metrics[1].price_per_unit}/{metrics[1].unit})"
            if len(metrics) > 1
            else "Silver (190000.00/10)"
        )
        composition_row_2 = [""] * len(headers)
        composition_row_2[18] = second_metric
        composition_row_2[19] = 3
        for col, value in enumerate(composition_row_2, 1):
            ws.cell(row=3, column=col, value=value)

        def make_variant_row(opt1_val, opt2_val, v_price, v_stock, v_image):
            row = [""] * len(headers)
            row[21] = opt1_val
            row[23] = opt2_val
            row[26] = v_price
            row[27] = v_stock
            row[28] = v_image
            return row

        variant_rows = [
            make_variant_row("Blue", "M", 85, 15, "variant_image_url"),
            make_variant_row("Blue", "L", 90, 20, "variant_image_url_2"),
            make_variant_row("Red", "S", 85, 12, "variant_image_url_3"),
        ]

        for row_num, variant_data in enumerate(variant_rows, 4):
            for col, value in enumerate(variant_data, 1):
                ws.cell(row=row_num, column=col, value=value)

        self._add_data_validations(wb, ws, categories, subcategories, metrics)
        self._adjust_column_widths(ws)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return FileResponse(
            buffer,
            as_attachment=True,
            filename="product_bulk_upload_template.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _add_data_validations(self, wb, ws, categories, subcategories, metrics):
        if "_dropdown_data" in wb.sheetnames:
            dropdown_ws = wb["_dropdown_data"]
        else:
            dropdown_ws = wb.create_sheet("_dropdown_data")
        dropdown_ws.sheet_state = "hidden"

        def add_bool_validation(cell_range):
            dv = DataValidation(
                type="list",
                formula1='"TRUE,FALSE"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid value",
                error="Please select TRUE or FALSE",
            )
            ws.add_data_validation(dv)
            dv.add(cell_range)

        add_bool_validation("E2:E1048576")
        add_bool_validation("M2:M1048576")
        add_bool_validation("N2:N1048576")
        add_bool_validation("P2:P1048576")
        add_bool_validation("R2:R1048576")

        status_dv = DataValidation(
            type="list",
            formula1='"active,draft,archived"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid status",
            error="Please select from: active, draft, archived",
        )
        ws.add_data_validation(status_dv)
        status_dv.add("O2:O1048576")

        if categories.exists():
            cat_names = [c.name for c in categories]
            for i, name in enumerate(cat_names, 1):
                dropdown_ws.cell(row=i, column=1, value=name)
            cat_dv = DataValidation(
                type="list",
                formula1=f"'_dropdown_data'!$A$1:$A${len(cat_names)}",
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid category",
                error="Please select from available categories",
            )
            ws.add_data_validation(cat_dv)
            cat_dv.add("K2:K1048576")

        if subcategories.exists():
            sub_names = [f"{s.name} ({s.category.name})" for s in subcategories]
            for i, name in enumerate(sub_names, 1):
                dropdown_ws.cell(row=i, column=2, value=name)
            sub_dv = DataValidation(
                type="list",
                formula1=f"'_dropdown_data'!$B$1:$B${len(sub_names)}",
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid subcategory",
                error="Please select from available subcategories",
            )
            ws.add_data_validation(sub_dv)
            sub_dv.add("L2:L1048576")

        if metrics:
            metric_display_names = [
                f"{m.name} ({m.price_per_unit}/{m.unit})" for m in metrics
            ]
            for i, name in enumerate(metric_display_names, 1):
                dropdown_ws.cell(row=i, column=3, value=name)
            comp_dv = DataValidation(
                type="list",
                formula1=f"'_dropdown_data'!$C$1:$C${len(metric_display_names)}",
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid composition",
                error="Please select a valid metric from the dropdown",
            )
            ws.add_data_validation(comp_dv)
            comp_dv.add("S2:S1048576")

        qty_dv = DataValidation(
            type="decimal",
            operator="greaterThan",
            formula1="0",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid quantity",
            error="Enter a numeric quantity greater than 0",
        )
        ws.add_data_validation(qty_dv)
        qty_dv.add("T2:T1048576")

    def _adjust_column_widths(self, ws):
        column_widths = {
            "A": 15,
            "B": 20,
            "C": 10,
            "D": 15,
            "E": 12,
            "F": 10,
            "G": 10,
            "H": 20,
            "I": 20,
            "J": 20,
            "K": 15,
            "L": 25,
            "M": 12,
            "N": 12,
            "O": 10,
            "P": 20,
            "Q": 20,
            "R": 20,
            "S": 30,
            "T": 12,
            "U": 15,
            "V": 15,
            "W": 15,
            "X": 15,
            "Y": 15,
            "Z": 15,
            "AA": 15,
            "AB": 15,
            "AC": 20,
            "AD": 20,
            "AE": 25,
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width


# ─── Bulk Upload ──────────────────────────────────────────────────────────────


class BulkProductUploadView(APIView):
    serializer_class = BulkUploadSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [TenantJWTAuthentication]

    def post(self, request):
        serializer = BulkUploadSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        file = serializer.validated_data["file"]
        zip_file = serializer.validated_data.get("zip_file")

        try:
            images_from_zip = {}
            if zip_file:
                images_from_zip = extract_images_from_zip(zip_file)

            if file.name.endswith(".csv"):
                df = pd.read_csv(file)
            elif file.name.endswith((".xls", ".xlsx")):
                df = pd.read_excel(file)
            else:
                return Response(
                    {"error": "Unsupported file format. Please upload .csv or .xlsx"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            df.columns = [col.strip().lower() for col in df.columns]
        except Exception as e:
            return Response(
                {"error": f"Invalid file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        all_metrics = list(
            PricingMetric.objects.only("id", "name", "price_per_unit", "unit")
        )
        metric_col_map = {
            f"{m.name} ({m.price_per_unit}/{m.unit})".strip().lower(): m
            for m in all_metrics
        }

        all_categories = {
            c.name.lower(): c for c in Category.objects.only("id", "name")
        }
        all_subcategories = {
            s.name.lower(): s for s in SubCategory.objects.only("id", "name")
        }

        created_products = {}
        current_product = None
        current_option_names = {}
        product_options = {}
        pending_compositions = {}

        for idx, row in df.iterrows():
            product_name = safe_value(row.get("name"))

            if product_name:
                if current_product and current_product.pk in pending_compositions:
                    self._save_compositions(
                        current_product, pending_compositions.pop(current_product.pk)
                    )

                if Product.objects.filter(name=product_name).exists():
                    current_product = None
                    current_option_names = {}
                    continue

                current_product_key = product_name

                if current_product_key not in created_products:
                    category_name = safe_value(row.get("category"))
                    sub_category_name = safe_value(row.get("subcategory"))

                    category = (
                        all_categories.get(category_name.lower())
                        if category_name
                        else None
                    )
                    sub_category = (
                        all_subcategories.get(sub_category_name.lower())
                        if sub_category_name
                        else None
                    )

                    raw_dynamic = safe_value(row.get("use_dynamic_pricing"), False)
                    if isinstance(raw_dynamic, str):
                        use_dynamic_pricing = raw_dynamic.strip().upper() == "TRUE"
                    else:
                        use_dynamic_pricing = bool(raw_dynamic)

                    base_making_charge = safe_value(row.get("base_making_charge"), 0.00)
                    try:
                        base_making_charge = (
                            float(base_making_charge)
                            if base_making_charge is not None
                            else 0.00
                        )
                    except (ValueError, TypeError):
                        base_making_charge = 0.00

                    raw_custom_image = safe_value(row.get("require_custom_image"), False)
                    if isinstance(raw_custom_image, str):
                        require_custom_image = raw_custom_image.strip().upper() == "TRUE"
                    else:
                        require_custom_image = bool(raw_custom_image)

                    product_data = {
                        "name": current_product_key,
                        "description": safe_value(row.get("description"), ""),
                        "price": safe_value(row.get("price"), 0),
                        "market_price": safe_value(row.get("market_price")),
                        "track_stock": safe_value(row.get("track_stock"), True),
                        "stock": safe_value(row.get("stock"), 0),
                        "weight": safe_value(row.get("weight")),
                        "thumbnail_alt_description": safe_value(
                            row.get(
                                "thumbnail_image_alt_description",
                                row.get("thumbnail_image_aly_description"),
                            ),
                            "",
                        ),
                        "category": category,
                        "sub_category": sub_category,
                        "is_popular": safe_value(row.get("is_popular"), False),
                        "is_featured": safe_value(row.get("is_featured"), False),
                        "status": safe_value(row.get("status"), "active"),
                        "use_dynamic_pricing": use_dynamic_pricing,
                        "base_making_charge": base_making_charge,
                        "require_custom_image": require_custom_image,
                        "meta_title": safe_value(
                            row.get("meta_title", row.get("meta title")), ""
                        ),
                        "meta_description": safe_value(
                            row.get("meta_description", row.get("meta description")), ""
                        ),
                    }

                    product_data = {
                        k: v for k, v in product_data.items() if v is not None
                    }
                    product = Product(**product_data)

                    thumb_val = safe_value(row.get("thumbnail_image"))
                    if thumb_val:
                        if not str(thumb_val).startswith(("http://", "https://")):
                            thumb_file = process_image_field(thumb_val, images_from_zip)
                            if thumb_file:
                                product.thumbnail_image.save(
                                    thumb_file.name, thumb_file, save=False
                                )
                        else:
                            thumb_file, thumb_filename = download_image_from_url(
                                thumb_val, upload_to="product_images"
                            )
                            if thumb_file:
                                product.thumbnail_image.save(
                                    thumb_filename, thumb_file, save=False
                                )

                    try:
                        product.save()
                    except Exception as e:
                        return Response(
                            {"error": str(e)}, status=status.HTTP_400_BAD_REQUEST
                        )

                    images_val = safe_value(row.get("images"))
                    if images_val:
                        image_names = [
                            img.strip()
                            for img in str(images_val).split(",")
                            if img.strip()
                        ]
                        for img_name in image_names:
                            if not img_name.startswith(("http://", "https://")):
                                img_file = process_image_field(
                                    img_name, images_from_zip
                                )
                                if img_file:
                                    prod_image = ProductImage(product=product)
                                    prod_image.image.save(
                                        img_file.name, img_file, save=True
                                    )
                            else:
                                img_file, img_filename = download_image_from_url(
                                    img_name, upload_to="product_images"
                                )
                                if img_file:
                                    prod_image = ProductImage(product=product)
                                    prod_image.image.save(
                                        img_filename, img_file, save=True
                                    )

                    option_names = self._create_product_options(product, row)
                    product_options[current_product_key] = option_names

                    created_products[current_product_key] = {
                        "product": product,
                        "options_created": True,
                        "use_dynamic_pricing": use_dynamic_pricing,
                    }
                    pending_compositions[product.pk] = []

                current_product = created_products[current_product_key]["product"]
                current_option_names = product_options[current_product_key]

            if current_product:
                product_meta = next(
                    (
                        v
                        for v in created_products.values()
                        if v["product"].pk == current_product.pk
                    ),
                    None,
                )
                use_dynamic = (
                    product_meta["use_dynamic_pricing"] if product_meta else False
                )

                if use_dynamic:
                    composition_val = safe_value(row.get("composition"))
                    quantity_val = safe_value(row.get("quantity"))

                    if composition_val and quantity_val is not None:
                        normalised_comp = str(composition_val).strip().lower()
                        metric = metric_col_map.get(normalised_comp)
                        try:
                            quantity = float(quantity_val)
                        except (ValueError, TypeError):
                            quantity = None

                        if metric and quantity and quantity > 0:
                            pending_compositions.setdefault(
                                current_product.pk, []
                            ).append((metric, quantity))

                self._create_product_variant(
                    current_product, row, current_option_names, images_from_zip
                )

        if current_product and current_product.pk in pending_compositions:
            self._save_compositions(
                current_product, pending_compositions.pop(current_product.pk)
            )

        return Response({
            "success": True,
            "message": f"Successfully processed {len(created_products)} products with variants",
        })

    def _save_compositions(self, product, composition_list):
        ProductComposition.objects.filter(product=product).delete()
        for metric, quantity in composition_list:
            ProductComposition.objects.create(
                product=product,
                metric=metric,
                quantity=quantity,
            )

    def _create_product_options(self, product, row):
        ProductOption.objects.filter(product=product).delete()
        option_names = {}
        option_columns = {}

        for col_name in row.index:
            col_str = str(col_name).lower()
            match = re.search(r"option(\d+)\s*name", col_str)
            if match:
                option_columns[int(match.group(1))] = col_name

        for option_num, col_name in sorted(option_columns.items()):
            option_name = safe_value(row.get(col_name))
            values_col_name = f"option{option_num} values"
            option_values_str = safe_value(row.get(values_col_name))

            if option_name:
                option_names[option_num] = option_name
                option, _ = ProductOption.objects.get_or_create(
                    product=product, name=option_name
                )
                if option_values_str:
                    ProductOptionValue.objects.get_or_create(
                        option=option, value=option_values_str
                    )

        return option_names

    def _create_product_variant(self, product, row, option_names, images_from_zip=None):
        variant_price = safe_value(row.get("variant price"))
        variant_stock = safe_value(row.get("variant stock"))
        variant_image_val = safe_value(row.get("variant image"))

        if variant_price is None and variant_stock is None:
            return None

        variant = ProductVariant(
            product=product,
            price=variant_price if variant_price is not None else product.price,
            stock=variant_stock if variant_stock is not None else (product.stock or 0),
        )

        if variant_image_val:
            if images_from_zip and not str(variant_image_val).startswith((
                "http://",
                "https://",
            )):
                variant_file = process_image_field(variant_image_val, images_from_zip)
                if variant_file:
                    variant.image.save(variant_file.name, variant_file, save=False)
            else:
                variant_file, variant_filename = download_image_from_url(
                    variant_image_val, upload_to="variant_images"
                )
                if variant_file:
                    variant.image.save(variant_filename, variant_file, save=False)

        variant.save()

        for option_num, option_name in option_names.items():
            values_col_name = f"option{option_num} values"
            option_value_str = safe_value(row.get(values_col_name))
            if not option_value_str:
                continue

            option = ProductOption.objects.filter(
                product=product, name=option_name
            ).first()

            if option:
                option_value, _ = ProductOptionValue.objects.get_or_create(
                    option=option, value=option_value_str
                )
                variant.option_values.add(option_value)

        return variant


# ─── Offer ────────────────────────────────────────────────────────────────────


OFFER_QS = Offer.objects.prefetch_related(
    "products", "categories", "sub_categories"
).order_by("-created_at")


class OfferListCreateView(generics.ListCreateAPIView):
    queryset = OFFER_QS
    pagination_class = CustomPagination

    def get_serializer_class(self):
        if self.request.method in permissions.SAFE_METHODS:
            return OfferSerializer
        return OfferWriteSerializer


class OfferRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = OFFER_QS

    def get_serializer_class(self):
        if self.request.method in permissions.SAFE_METHODS:
            return OfferSerializer
        return OfferWriteSerializer


class OfferProductListView(generics.ListAPIView):
    queryset = PRODUCT_LIST_QS
    serializer_class = ProductSerializer
    pagination_class = ProductPagination
    filter_backends = [
        django_filters.DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_class = ProductFilterSet
    ordering_fields = ["created_at", "price", "is_popular", "average_rating"]
    search_fields = ["name"]

    def get_serializer_class(self):
        site_config = SiteConfig.get_solo()
        if site_config.use_product_variant:
            return UnifiedProductListingSerializer
        return ProductSmallSerializer

    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)
        ordering = self.request.query_params.get("ordering", "")
        if ordering.lstrip("-") == "price":
            direction = "-" if ordering.startswith("-") else ""
            queryset = queryset.order_by(f"{direction}computed_final_price")
        return queryset

    def list(self, request, *args, **kwargs):
        site_config = SiteConfig.get_solo()

        from django.db.models import Q
        from django.utils import timezone

        now = timezone.now()
        active_offers = Offer.objects.filter(
            is_active=True, start_date__lte=now, end_date__gte=now
        )

        has_offer_products = False
        q_prod = Q()
        q_var = Q()
        if active_offers.exists():
            category_ids = list(
                filter(None, active_offers.values_list("categories", flat=True))
            )
            sub_category_ids = list(
                filter(None, active_offers.values_list("sub_categories", flat=True))
            )
            product_ids = list(
                filter(None, active_offers.values_list("products", flat=True))
            )

            if category_ids:
                q_prod |= Q(category_id__in=category_ids)
                q_var |= Q(product__category_id__in=category_ids)
            if sub_category_ids:
                q_prod |= Q(sub_category_id__in=sub_category_ids)
                q_var |= Q(product__sub_category_id__in=sub_category_ids)
            if product_ids:
                q_prod |= Q(id__in=product_ids)
                q_var |= Q(product_id__in=product_ids)

            if (
                category_ids or sub_category_ids or product_ids
            ) and Product.objects.filter(q_prod).exists():
                has_offer_products = True

        if not site_config.use_product_variant:
            queryset = self.get_queryset()
            if has_offer_products:
                queryset = queryset.filter(q_prod).distinct()

            queryset = self.filter_queryset(queryset)

            if has_offer_products:
                max_price = float(
                    Product.objects
                    .filter(q_prod)
                    .annotate(computed_final_price=FINAL_PRICE_ANNOTATION)
                    .aggregate(m=Max("computed_final_price"))["m"]
                    or 0.0
                )
            else:
                max_price = float(
                    Product.objects.annotate(
                        computed_final_price=FINAL_PRICE_ANNOTATION
                    ).aggregate(m=Max("computed_final_price"))["m"]
                    or 0.0
                )

            page = self.paginate_queryset(queryset)
            if page is not None:
                if self.paginator is not None:
                    self.paginator.max_price = max_price
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)

        # ── variant mode ────────────────────────────────────────────────────────
        self.filter_backends = []

        params = request.GET.copy()

        from distutils.util import strtobool

        if "is_popular" in params:
            try:
                params["is_popular"] = strtobool(str(params["is_popular"]))
            except ValueError:
                params.pop("is_popular")

        variant_qs = PRODUCT_VARIANT_QS
        if has_offer_products:
            variant_qs = variant_qs.filter(q_var).distinct()

        variant_filter = ProductVariantFilterSet(params, queryset=variant_qs)
        variant_qs = variant_filter.qs

        search_filter = filters.SearchFilter()
        ordering_filter = filters.OrderingFilter()

        orig_search_fields = self.search_fields
        self.search_fields = ["product__name", "option_values__value"]

        variant_qs = search_filter.filter_queryset(request, variant_qs, self)
        variant_qs = ordering_filter.filter_queryset(request, variant_qs, self)

        self.search_fields = orig_search_fields

        product_qs = (
            Product.objects
            .filter(variants__isnull=True)
            .select_related("category", "sub_category")
            .prefetch_related(
                Prefetch(
                    "images",
                    queryset=ProductImage.objects.only("id", "product_id", "image"),
                ),
            )
            .annotate(
                average_rating=Avg("productreview__rating"),
                computed_final_price=FINAL_PRICE_ANNOTATION,
            )
            .only(
                "id",
                "name",
                "slug",
                "price",
                "market_price",
                "stock",
                "thumbnail_image",
                "thumbnail_alt_description",
                "category_id",
                "sub_category_id",
                "is_popular",
                "is_featured",
                "fast_shipping",
                "warranty",
                "use_dynamic_pricing",
                "base_making_charge",
                "created_at",
                "updated_at",
            )
        )
        if has_offer_products:
            product_qs = product_qs.filter(q_prod).distinct()

        product_filter = ProductFilterSet(params, queryset=product_qs)
        product_qs = product_filter.qs
        product_qs = search_filter.filter_queryset(request, product_qs, self)
        product_qs = ordering_filter.filter_queryset(request, product_qs, self)

        # ── ordering-aware Python sort for the combined list ──────────────────
        ordering = request.query_params.get("ordering", "-created_at")
        sort_field = ordering.lstrip("-")
        sort_reverse = ordering.startswith("-")

        def _sort_key(obj):
            if sort_field == "price":
                if isinstance(obj, ProductVariant):
                    return float(obj.price or 0)
                if (
                    hasattr(obj, "computed_final_price")
                    and obj.computed_final_price is not None
                ):
                    return float(obj.computed_final_price)
                return float(obj.final_price or 0)
            if sort_field == "average_rating":
                return float(getattr(obj, "average_rating", None) or 0)
            if sort_field == "is_popular":
                val = getattr(obj, "is_popular", False)
                if isinstance(obj, ProductVariant):
                    val = getattr(obj.product, "is_popular", False)
                return 0 if val else 1
            return getattr(obj, "created_at", None) or 0

        combined = sorted(
            chain(variant_qs, product_qs),
            key=_sort_key,
            reverse=sort_reverse,
        )

        if has_offer_products:
            variant_max = float(
                ProductVariant.objects.filter(q_var).aggregate(m=Max("price"))["m"]
                or 0.0
            )
            standalone_product_max = float(
                Product.objects
                .filter(q_prod, variants__isnull=True)
                .annotate(computed_final_price=FINAL_PRICE_ANNOTATION)
                .aggregate(m=Max("computed_final_price"))["m"]
                or 0.0
            )
            fallback_max = float(
                Product.objects.filter(
                    q_prod,
                    variants__isnull=False,
                    variants__price__isnull=True,
                ).aggregate(m=Max("price"))["m"]
                or 0.0
            )
        else:
            variant_max = float(
                ProductVariant.objects.aggregate(m=Max("price"))["m"] or 0.0
            )
            standalone_product_max = float(
                Product.objects
                .filter(variants__isnull=True)
                .annotate(computed_final_price=FINAL_PRICE_ANNOTATION)
                .aggregate(m=Max("computed_final_price"))["m"]
                or 0.0
            )
            fallback_max = float(
                Product.objects.filter(
                    variants__isnull=False,
                    variants__price__isnull=True,
                ).aggregate(m=Max("price"))["m"]
                or 0.0
            )
        max_price = max(variant_max, standalone_product_max, fallback_max)

        page = self.paginate_queryset(combined)
        if page is not None:
            if self.paginator is not None:
                self.paginator.max_price = max_price
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(combined, many=True)
        return Response(serializer.data)
